#!/usr/bin/env python3
"""
generate_workplan.py — Bellwether Workplan Generator

Reads a workplan.json file produced by the bw-workplan Claude Skill and fills in
a copy of the Bellwether workplan template, producing a formatted .xlsx file.

Usage:
    python generate_workplan.py workplan.json output.xlsx
"""

import json
import os
import sys
import subprocess
from datetime import datetime, date, timedelta
from io import SEEK_SET

# ── Auto-install openpyxl ─────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("Installing openpyxl...")
    subprocess.check_call(
        [sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter, column_index_from_string


# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(argb):
    return PatternFill("solid", fgColor=argb)


def font(bold=False, color="FF000000", size=10, name="Aptos"):
    return Font(bold=bold, color=color, size=size, name=name)


# ── Constants ─────────────────────────────────────────────────────────────────
PLUM    = "FF6D1E4A"
ORANGE  = "FFFFB653"
GRAY    = "FFD9D9D9"
GRAY_FONT = "FF808080"

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, "assets", "workplan-template.xlsx")

# Max weeks the Project Timeline template supports (cols B–Z = 25 columns)
MAX_WEEKS = 25

# Cover Sheet cell positions
CS_TITLE_CELL     = "A1"
CS_COST_CODE_CELL = "B2"
CS_TEAM_START_ROW = 8   # A8:A12 = Team Members 1–5
CS_PHASE_START_ROW = 15  # A15:A20 = Phase 0–5

# Project Timeline row positions
TL_INTERNAL_ROW    = 2
TL_EXTERNAL_ROW    = 3
TL_PHASE_ROW       = 4
TL_ACTIVITIES_ROW  = 5
TL_DEADLINES_ROW   = 6
TL_PTO_ROW         = 7
TL_TEAM_START_ROW  = 10  # A10:A14 = Team Members 1–5

# Detailed Workplan
WP_HEADER_ROW = 1
WP_DATA_START  = 2


# ── Date helpers ──────────────────────────────────────────────────────────────
def parse_date(s):
    """Parse YYYY-MM-DD string to date."""
    return datetime.strptime(s, "%Y-%m-%d").date()


def week_of(d):
    """Return the Monday of the week containing date d."""
    return d - timedelta(days=d.weekday())


# ── Validation ────────────────────────────────────────────────────────────────
def validate_json(data):
    """Basic sanity checks on the workplan JSON."""
    errors = []
    required_top = ["metadata", "project", "calendar", "tasks", "milestones", "timeline"]
    for key in required_top:
        if key not in data:
            errors.append(f"Missing top-level key: '{key}'")

    if "calendar" in data:
        for f in ["start_date", "end_date"]:
            if f not in data["calendar"]:
                errors.append(f"Missing calendar.{f}")

    if "tasks" in data:
        seen_ids = set()
        for t in data["tasks"]:
            tid = t.get("id", "?")
            if tid in seen_ids:
                errors.append(f"Duplicate task id: {tid}")
            seen_ids.add(tid)

    if errors:
        for e in errors:
            print(f"  ERROR: {e}", file=sys.stderr)
        sys.exit(1)


# ── Sheet filling functions ───────────────────────────────────────────────────
def fill_cover_sheet(ws, data):
    """Populate the Cover Sheet with project metadata."""
    project  = data["project"]
    calendar = data["calendar"]

    # Project title
    ws[CS_TITLE_CELL] = project.get("name", "Untitled Project")

    # Cost code
    ws[CS_COST_CODE_CELL] = project.get("cost_code") or ""

    # Team members (A8:A12)
    members = project.get("team_members", [])
    for i in range(5):
        val = members[i] if i < len(members) else ""
        ws.cell(row=CS_TEAM_START_ROW + i, column=1).value = val or ""

    # Phase names (A15:A20) — extract unique phases in order from tasks
    phases_seen = []
    for task in data.get("tasks", []):
        phase = task.get("phase", "")
        if phase and phase not in phases_seen:
            phases_seen.append(phase)

    for i in range(6):
        ws.cell(row=CS_PHASE_START_ROW + i, column=1).value = (
            phases_seen[i] if i < len(phases_seen) else ""
        )


def fill_timeline(ws, data):
    """Populate the Project Timeline sheet."""
    timeline = data.get("timeline", [])
    project  = data["project"]
    members  = project.get("team_members", [])

    if not timeline:
        print("  Warning: timeline[] is empty; Project Timeline sheet will be blank.")
        return

    n_weeks = min(len(timeline), MAX_WEEKS)
    if len(timeline) > MAX_WEEKS:
        print(
            f"  Warning: Project spans {len(timeline)} weeks; "
            f"truncating to {MAX_WEEKS} (template limit)."
        )

    # Write team member names to A10:A14
    for i in range(5):
        ws.cell(row=TL_TEAM_START_ROW + i, column=1).value = (
            members[i] if i < len(members) else ""
        )

    # Week header row (row 1, cols B+) and data rows
    for week_idx, week in enumerate(timeline[:n_weeks]):
        col = week_idx + 2  # col B = index 2

        # Week date header
        ws.cell(row=1, column=col).value = week.get("week_display", "")

        # Activity rows
        ws.cell(row=TL_INTERNAL_ROW,   column=col).value = week.get("internal_meetings", "")
        ws.cell(row=TL_EXTERNAL_ROW,   column=col).value = week.get("external_meetings", "")
        ws.cell(row=TL_PHASE_ROW,      column=col).value = week.get("phase", "")
        ws.cell(row=TL_ACTIVITIES_ROW, column=col).value = week.get("key_activities", "")
        ws.cell(row=TL_DEADLINES_ROW,  column=col).value = week.get("deadlines_milestones", "")
        ws.cell(row=TL_PTO_ROW,        column=col).value = week.get("pto", "")

        # Team capacity rows
        team_capacity = week.get("team_capacity", [])
        for member_idx in range(5):
            ws.cell(
                row=TL_TEAM_START_ROW + member_idx,
                column=col,
            ).value = (
                team_capacity[member_idx] if member_idx < len(team_capacity) else ""
            )

    # Hide unused week columns (col n_weeks+2 through Z)
    for col_idx in range(n_weeks + 2, MAX_WEEKS + 2):
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True


def fill_workplan(ws, data):
    """Populate the Detailed Workplan sheet."""
    tasks    = data.get("tasks", [])
    calendar = data["calendar"]

    # Delete existing data rows
    if ws.max_row > WP_HEADER_ROW:
        ws.delete_rows(WP_DATA_START, ws.max_row - WP_HEADER_ROW)

    # Feasibility warning banner
    feasibility_flag = calendar.get("feasibility_flag", "ok")
    feasibility_note = calendar.get("feasibility_note", "")
    row_offset = 0

    if feasibility_flag != "ok":
        ws.insert_rows(WP_DATA_START)
        warning_cell = ws.cell(row=WP_DATA_START, column=3)  # Task column
        warning_cell.value = f"⚠ {feasibility_note}"
        warning_cell.font  = Font(bold=True, color="FF000000", size=10, name="Aptos")
        for col in range(1, 9):
            ws.cell(row=WP_DATA_START, column=col).fill = fill(ORANGE)
        row_offset = 1

    # Write task rows
    for task_idx, task in enumerate(tasks):
        row_num = WP_DATA_START + row_offset + task_idx

        ws.cell(row=row_num, column=1).value = task.get("phase", "")
        ws.cell(row=row_num, column=2).value = task.get("workstream") or ""
        ws.cell(row=row_num, column=3).value = task.get("task", "")
        ws.cell(row=row_num, column=4).value = task.get("owner") or ""
        ws.cell(row=row_num, column=5).value = "Not started"
        ws.cell(row=row_num, column=6).value = task.get("start_date_display", "")
        ws.cell(row=row_num, column=7).value = task.get("deadline_display", "")
        ws.cell(row=row_num, column=8).value = task.get("notes") or ""

        # Alternating row fill
        row_fill = fill("FFE2D2DB") if task_idx % 2 == 0 else fill("FFFFFFFF")
        for col in range(1, 9):
            c = ws.cell(row=row_num, column=col)
            c.fill = row_fill
            c.font = Font(size=10, name="Aptos", color="FF595959")
            if col in [3, 8]:
                c.alignment = Alignment(wrap_text=True, vertical="center")
            else:
                c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_num].height = 18

    # Add conditional formatting: Status = "Completed" → gray row
    last_data_row = WP_DATA_START + row_offset + len(tasks) + 50  # buffer
    cf_range = f"A{WP_DATA_START}:H{last_data_row}"
    gray_fill_obj = PatternFill("solid", fgColor="FFD9D9D9")
    gray_font_obj = Font(color="FF808080", size=10, name="Aptos")
    ws.conditional_formatting.add(
        cf_range,
        FormulaRule(
            formula=[f'$E{WP_DATA_START}="Completed"'],
            fill=gray_fill_obj,
            font=gray_font_obj,
        ),
    )

    # Replace/expand data validation to cover E2:E1000
    dv = DataValidation(
        type="list",
        formula1='"Not started,To-Do,In progress,Completed"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = f"E{WP_DATA_START}:E1000"
    # Remove old validations to avoid duplicates
    ws.data_validations.dataValidation = []
    ws.add_data_validation(dv)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) != 3:
        print("Usage: python generate_workplan.py workplan.json output.xlsx", file=sys.stderr)
        sys.exit(1)

    json_path   = sys.argv[1]
    output_path = sys.argv[2]

    # Load JSON
    if not os.path.exists(json_path):
        print(f"Error: {json_path} not found", file=sys.stderr)
        sys.exit(1)
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    print(f"Loaded: {json_path}")
    validate_json(data)

    # Verify template
    if not os.path.exists(TEMPLATE_PATH):
        print(f"Error: template not found at {TEMPLATE_PATH}", file=sys.stderr)
        sys.exit(1)

    # Open template
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    cover    = wb["Cover Sheet"]
    timeline = wb["Project Timeline"]
    workplan = wb["Detailed Workplan"]

    # Fill sheets
    print("Filling Cover Sheet...")
    fill_cover_sheet(cover, data)

    print("Filling Project Timeline...")
    fill_timeline(timeline, data)

    n_weeks = min(len(data.get("timeline", [])), MAX_WEEKS)
    print(f"  {n_weeks} week columns written")

    print("Filling Detailed Workplan...")
    fill_workplan(workplan, data)

    n_tasks = len(data.get("tasks", []))
    print(f"  {n_tasks} task rows written")

    # Save
    wb.save(output_path)
    size = os.path.getsize(output_path)

    # Summary
    calendar = data["calendar"]
    feasibility = calendar.get("feasibility_flag", "ok").upper()
    avail   = calendar.get("available_business_days", "?")
    est     = calendar.get("estimated_total_days", "?")
    note    = calendar.get("feasibility_note", "")

    print()
    print(f"Template: {os.path.basename(TEMPLATE_PATH)}")
    print(f"Project:  {data['project'].get('name', 'Untitled')}")
    print(f"Timeline: {calendar.get('start_date', '?')} → {calendar.get('end_date', '?')}")
    print(f"          ({n_weeks} weeks, {avail} available business days)")
    print(f"Feasibility: {feasibility} ({est} estimated days / {avail} available)")
    if feasibility != "OK" and note:
        print(f"  Note: {note}")
    print(f"Tasks:    {n_tasks} rows written")
    print(f"Output:   {output_path} ({size:,} bytes)")


if __name__ == "__main__":
    main()
